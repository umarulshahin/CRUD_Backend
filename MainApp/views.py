from django.shortcuts import render,HttpResponse
from rest_framework.response import Response
from rest_framework.decorators import api_view,permission_classes
from rest_framework.permissions import IsAuthenticated
from .serializer import RegisterSerializer,UserSerializer,ImageUploadSerializer
from .models import *
from rest_framework_simplejwt.serializers import TokenObtainPairSerializer
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.views import APIView
from rest_framework_simplejwt.authentication import JWTAuthentication
from openpyxl import Workbook
from openpyxl.styles import *
from datetime import datetime
from openpyxl.utils import get_column_letter
from rest_framework import status

    
# Create your views here.


@api_view(['GET'])  
@permission_classes([IsAuthenticated])  
def Dashboard(request):
    if request.method == 'GET':
        user = request.user
        
        data=CustomUser.objects.filter(email=user)
        serializer=UserSerializer(data,many=True)
           
        return Response(serializer.data)

    return Response({'error': "Method not allowed"})
   

@api_view(['POST'])
def Sign_Up(request):
    
   if request.method == 'POST':
       nested_data = request.data.get('data', {})
       data = request.data
       print(data,'register data')
       serializer=RegisterSerializer(data=data['data'])  
      
       if serializer.is_valid():
           serializer.save()
           print(serializer.data,'serializer data')
           return Response(serializer.data)  
       
       return Response({"error":serializer.errors},status=status.HTTP_400_BAD_REQUEST)
    
@api_view(["PATCH"])
@permission_classes([IsAuthenticated])    
def Image_Upload(request):
    
    if request.method == "PATCH":
        image = request.FILES.get('image')
        user_email = request.user.email
        print(image, "image")
        if not image:
            return Response({"error": "Image file is required"}, status=status.HTTP_400_BAD_REQUEST)
        data = {"image": image, "user": user_email}

        serializer = ImageUploadSerializer(data=data)
        if serializer.is_valid():
            serializer.save()
            return Response({"Success": "Image uploaded successfully"}, status=status.HTTP_200_OK)
        return Response({"error": serializer.errors}, status=status.HTTP_400_BAD_REQUEST)

class MyTokenObtainPairSerializer(TokenObtainPairSerializer):
    
    @classmethod
    def get_token(cls, user):
        token = super().get_token(user)

        token['username'] = user.username
        token["role"]=user.is_superuser
        
        return token

class MyTokenobtainedPairView(TokenObtainPairView):
    serializer_class=MyTokenObtainPairSerializer
    
class Update_User(APIView):
    
    permission_classes([IsAuthenticated])
    def patch(self,request):
        
        id=request.data.get("id")
        
        obj=CustomUser.objects.filter(id=id).first()
        if not obj:
            return Response({"error": "User not found"})
        print(id,'user id')
        print(request.data)
        serializer=UserSerializer(obj,data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()
            return Response(serializer.data)
        return Response({"error":serializer.errors})
    
    
@api_view(["POST"])
def UserdataExel(request):
    id = request.data
    print(id)
    user = CustomUser.objects.get(id=id)
    profile_url = request.build_absolute_uri(user.profile.url)
    print(profile_url,'profile url')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Users Data'
    
    # Define headers first
    headers = ['Id', 'First Name', 'Last Name', 'Email', 'Phone', 'Date Joined','Profile']
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Now merge cells for title after headers are written
    ws.merge_cells('A1:F1')  # Updated to include all columns
    ws['A1'] = 'User Details Report'
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].font = Font(bold=True, size=14)
    
    # Convert date
    date_joined = timezone.localtime(user.date_joined).replace(tzinfo=None)
    
    # User data
    user_data = [
        user.id,
        user.username,
        user.last_name,
        user.email,
        user.phone,
        date_joined,
        profile_url
    ]
    
    # Write user data
    for col, value in enumerate(user_data, 1):
        cell = ws.cell(row=3, column=col, value=value)
        cell.alignment = Alignment(horizontal='left')
    
    # Auto-adjust column widths - fixed version
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))  # Start with header length
        
        # Check data cell in this column
        data_value = str(user_data[col_idx - 1])
        max_length = max(max_length, len(data_value))
        
        # Set column width with padding
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = max_length + 2
    
    # Apply borders to all cells in use
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=3):
        for cell in row:
            cell.border = thin_border
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=user_details.xlsx'
    
    wb.save(response)
    
    if hasattr(response, 'seek'):
        response.seek(0)
        
    return response